import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

planilha_clientes = openpyxl.load_workbook("dados_clientes.xlsx")
pagina = planilha_clientes['Sheet1']

driver = webdriver.Chrome()
driver.get("https://consultcpf-devaprender.netlify.app/")
sleep(5)

for i in pagina.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = i

    pesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
    pesquisa.clear()
    sleep(1)
    pesquisa.send_keys(cpf)
    consultar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(1)
    consultar.click()
    sleep(4)
    status = driver.find_element(By.XPATH,"//span[@id='statusLabel']").text
    if status == "em dia":
        data_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentDate']").text
        data_pagamento_limpo = data_pagamento.split()[3]
        metodo_pagamento = driver.find_element(By.XPATH,"//p[@id='paymentMethod']").text
        metodo_pagamento_limpo = metodo_pagamento.split()[3]

        planilha_fechamento = openpyxl.load_workbook("planilha_fechamento.xlsx")
        pagina_fechamento = planilha_fechamento['Sheet1']

        pagina_fechamento.append([nome,valor,cpf,vencimento,status,data_pagamento_limpo,metodo_pagamento_limpo])
        planilha_fechamento.save("planilha_fechamento.xlsx")


    else:
        status = "pendente"
        planilha_fechamento = openpyxl.load_workbook("planilha_fechamento.xlsx")
        pagina_fechamento = planilha_fechamento['Sheet1']
        pagina_fechamento.append([nome,valor,cpf,vencimento,status])
        planilha_fechamento.save("planilha_fechamento.xlsx")



   